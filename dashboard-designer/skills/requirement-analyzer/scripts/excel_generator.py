"""
数据模型Excel生成脚本
生成包含7个Sheet的数据模型设计Excel文件
"""

import sys
import argparse
from pathlib import Path
from datetime import datetime

# 添加共享脚本路径
sys.path.insert(0, str(Path(__file__).parent.parent.parent / 'shared' / 'scripts'))

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("错误: openpyxl未安装，请运行: pip install openpyxl")
    sys.exit(1)


# 样式定义
HEADER_FONT = Font(bold=True, size=11)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT_WHITE = Font(bold=True, size=11, color="FFFFFF")
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


def create_sheet_with_headers(wb, sheet_name: str, headers: list) -> None:
    """创建带表头的Sheet"""
    ws = wb.create_sheet(title=sheet_name)

    # 写入表头
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN_BORDER

    # 设置列宽
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18

    return ws


def add_row(ws, row_data: list, row_num: int) -> None:
    """添加一行数据"""
    for col, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_num, column=col, value=value)
        cell.border = THIN_BORDER
        cell.alignment = Alignment(vertical='center')


def generate_data_model_excel(project_name: str, output_path: str, data: dict = None) -> str:
    """
    生成数据模型设计Excel文件

    Args:
        project_name: 项目名称
        output_path: 输出目录路径
        data: 数据内容（可选，用于填充示例数据）

    Returns:
        生成的Excel文件路径
    """
    wb = Workbook()

    # 删除默认Sheet
    wb.remove(wb.active)

    # ========== Sheet 1: 页面结构 ==========
    ws1 = create_sheet_with_headers(wb, "页面结构", [
        "页面ID", "页面名称", "页面描述", "布局方式", "核心功能", "父页面"
    ])

    # 示例数据
    sample_pages = data.get("pages", [
        ["P001", "概览页", "展示核心KPI和趋势", "3行布局", "KPI展示、趋势分析", "-"],
        ["P002", "区域分析页", "区域销售分布分析", "左右分栏", "地图展示、区域对比", "P001"],
        ["P003", "产品分析页", "产品销售明细", "上下布局", "明细查询、导出", "P001"]
    ])

    for i, row in enumerate(sample_pages, 2):
        add_row(ws1, row, i)

    # ========== Sheet 2: 组件清单 ==========
    ws2 = create_sheet_with_headers(wb, "组件清单", [
        "组件ID", "所属页面", "组件名称", "组件类型", "位置说明", "交互说明"
    ])

    sample_components = data.get("components", [
        ["C001", "P001", "销售额KPI", "指标卡", "第一行左1", "点击跳转区域分析"],
        ["C002", "P001", "销售趋势图", "折线图", "第二行左侧", "支持缩放、拖动"],
        ["C003", "P001", "区域占比图", "饼图", "第二行右侧", "点击下钻"],
        ["C004", "P002", "区域销售地图", "地图", "左侧主区域", "支持区域选择"],
        ["C005", "P003", "产品明细表", "表格", "主区域", "支持排序、筛选"]
    ])

    for i, row in enumerate(sample_components, 2):
        add_row(ws2, row, i)

    # ========== Sheet 3: 数据模型-表结构 ==========
    ws3 = create_sheet_with_headers(wb, "数据模型-表结构", [
        "表ID", "表名称", "表英文名", "说明", "主键字段", "更新频率"
    ])

    sample_tables = data.get("tables", [
        ["T001", "月度销售汇总表", "monthly_sales_summary", "各事业部月度销售汇总", "month + business_unit", "T+1"],
        ["T002", "区域销售汇总表", "regional_sales_summary", "各区域销售汇总", "region_code + province", "T+1"],
        ["T003", "产品销售汇总表", "product_sales_summary", "产品销售汇总", "category_name + product_name", "T+1"]
    ])

    for i, row in enumerate(sample_tables, 2):
        add_row(ws3, row, i)

    # ========== Sheet 4: 数据模型-字段定义 ==========
    ws4 = create_sheet_with_headers(wb, "数据模型-字段定义", [
        "字段ID", "所属表ID", "字段名称", "字段英文名", "字段类型", "说明", "示例值"
    ])

    sample_fields = data.get("fields", [
        ["F001", "T001", "月份", "month", "VARCHAR(7)", "数据月份", "2024-01"],
        ["F002", "T001", "事业部", "business_unit", "VARCHAR(50)", "业务单元", "华东事业部"],
        ["F003", "T001", "销售额", "sales_amount", "DECIMAL(12,2)", "销售金额", "1285000.00"],
        ["F004", "T001", "同比增长率", "yoy_growth", "DECIMAL(5,2)", "同比增长", "12.35"],
        ["F005", "T002", "区域名称", "region_name", "VARCHAR(50)", "销售区域", "华东"],
        ["F006", "T002", "销售额", "sales_amount", "DECIMAL(12,2)", "销售金额", "568000.00"]
    ])

    for i, row in enumerate(sample_fields, 2):
        add_row(ws4, row, i)

    # ========== Sheet 5: 组件与数据模型映射 ==========
    ws5 = create_sheet_with_headers(wb, "组件与数据模型映射", [
        "映射ID", "组件ID", "组件名称", "数据表ID", "数据表名称", "使用字段", "筛选条件", "聚合方式"
    ])

    sample_mappings = data.get("mappings", [
        ["M001", "C001", "销售额KPI", "T001", "monthly_sales_summary", "sales_amount, yoy_growth", "当月", "SUM"],
        ["M002", "C002", "销售趋势图", "T001", "monthly_sales_summary", "month, sales_amount", "近6月", "按月GROUP BY"],
        ["M003", "C003", "区域占比图", "T002", "regional_sales_summary", "region_name, sales_amount", "当月", "按区域GROUP BY"]
    ])

    for i, row in enumerate(sample_mappings, 2):
        add_row(ws5, row, i)

    # ========== Sheet 6: 模型关联关系 ==========
    ws6 = create_sheet_with_headers(wb, "模型关联关系", [
        "关联ID", "主表ID", "主表名称", "从表ID", "从表名称", "关联字段(主表)", "关联字段(从表)", "关联说明"
    ])

    sample_relations = data.get("relations", [
        ["R001", "T001", "monthly_sales_summary", "T002", "regional_sales_summary", "business_unit", "region_name", "事业部与区域映射"],
        ["R002", "T001", "monthly_sales_summary", "T003", "product_sales_summary", "month", "month", "时间维度关联"]
    ])

    for i, row in enumerate(sample_relations, 2):
        add_row(ws6, row, i)

    # ========== Sheet 7: 维度定义 ==========
    ws7 = create_sheet_with_headers(wb, "维度定义", [
        "维度ID", "维度名称", "维度英文名", "数据类型", "层级结构", "支持筛选", "支持下钻", "说明"
    ])

    sample_dimensions = data.get("dimensions", [
        ["D001", "时间", "month", "VARCHAR(7)", "年-季-月-日", "是", "是", "主维度"],
        ["D002", "事业部", "business_unit", "VARCHAR(50)", "单层", "是", "否", "组织维度"],
        ["D003", "区域", "region_name", "VARCHAR(50)", "区域-省-市", "是", "是", "地理维度"]
    ])

    for i, row in enumerate(sample_dimensions, 2):
        add_row(ws7, row, i)

    # 保存文件
    output_dir = Path(output_path)
    output_dir.mkdir(parents=True, exist_ok=True)
    file_path = output_dir / f"{project_name}_数据模型设计.xlsx"
    wb.save(file_path)

    return str(file_path)


def main():
    parser = argparse.ArgumentParser(description='数据模型Excel生成工具')
    parser.add_argument('--project', '-p', required=True, help='项目名称')
    parser.add_argument('--output', '-o', default='output', help='输出目录路径')
    parser.add_argument('--data', '-d', help='数据JSON文件路径（可选）')

    args = parser.parse_args()

    # 加载数据（如果有）
    data = {}
    if args.data:
        import json
        with open(args.data, 'r', encoding='utf-8') as f:
            data = json.load(f)

    # 生成Excel
    file_path = generate_data_model_excel(args.project, args.output, data)
    print(f"Excel文件已生成: {file_path}")


if __name__ == '__main__':
    main()
